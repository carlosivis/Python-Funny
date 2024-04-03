import openpyxl
import random
import pandas as pd


def gen_sheet(filename, headers, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for col, header in enumerate(headers.keys(), start=1):
        sheet.cell(row=1, column=col, value=header)

    for row, point_data in enumerate(data, start=2):
        for col, key in enumerate(headers.keys(), start=1):
            sheet.cell(row=row, column=col, value=point_data[key])

    workbook.save(filename)


def gen_points_name(qtd):
    names: list[str] = []
    for name in range(1, qtd+1):
        names.append(f"Ponto{name}")
    return names


def gen_random_data(points, headers, qtd, start_date, end_date):
    data = []

    #for point in points:
    date_range = pd.date_range(start_date, end_date)
    date_iter = iter(date_range)
    for _ in range(qtd):
        try:
            date = next(date_iter)  # Get the next date from the iterator
        except StopIteration:  # If all dates have been used, raise an error
            raise ValueError("Quantity exceeds available dates in the specified time interval.")

        # point_data = {"Ponto": point, "Data": date.strftime("%d/%m/%Y %H:%M:%S")}
        point_data = {"Ponto": random.choice(points), "Data": date.strftime("%d/%m/%Y %H:%M:%S")}
        for key, limits in headers.items():
            if key != "Ponto" and limits is not None:
                min_limit, max_limit = limits
                value = round(random.uniform(min_limit, max_limit), 2)

                # If not the first point, interpolate with previous value
                if data:
                    previous_value = data[-1][key]
                    interpolated_value = (value + previous_value) / 2
                    value = interpolated_value

                point_data[key] = value
        data.append(point_data)

    # Modify % of the values in each column to be outliers
    for key, limits in headers.items():
        if key != "Ponto" and key != "Data" and limits is not None:
            num_points_to_modify = int(len(data) * 0.02)
            points_to_modify = random.sample(data, num_points_to_modify)

            for point in points_to_modify:
                min_limit, max_limit = limits

                outlier_direction = random.choice(["above", "below"])

                if outlier_direction == "above":
                    outlier_value = random.uniform(max_limit * 1.1, max_limit * 1.5)
                elif outlier_direction == "below":
                    outlier_value = random.uniform(min_limit * 0.1, min_limit * 0.5)
                else:
                    raise ValueError("Invalid outlier direction")

                point[key] = round(outlier_value, 2)
    return data


#Header can be changed to ur neccesity, try like: "header  name": (x,y) also: x is lower_limit and y is upper_limit
headers = {
    "Ponto": None,
    "Data": None,
    "Temperatura da agua": (22, 25),
    "pH": (6, 8),
    "Coliformes Termotolerantes": (2000, 2600),
    "Turbidez": (3, 5),
    "Oxigenio Dissolvido": (10, 15),
    "Demanda bioquímica de oxigenio": (5, 6),
    "Solidos Totais": (100, 110),
    "Fosforo Total":(2,4),
    "Nitrogenio Total":(2,4)
}

points = gen_points_name(10)
data = gen_random_data(points, headers, 500,"2020-01-01","2024-01-30")
gen_sheet("data_with_2%_outlier.xlsx", headers, data)
