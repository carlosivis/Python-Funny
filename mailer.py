import smtplib
import argparse
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from datetime import datetime


def config_service(host, pwd, use_local=True):
    if use_local:
        server = smtplib.SMTP('localhost', 1025)  # Escolha uma porta disponível
    else:
        server = smtplib.SMTP('smtp.gmail.com:587')
        server.starttls()
        server.login(host, pwd)
    return server


def send_mail(server, from_email, to, subject, body, attach):
    try:
        server.ehlo_or_helo_if_needed()
        # Restante do código...

        server.ehlo_or_helo_if_needed()  # Certifica-se de que a conexão está estabelecida

        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = to
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        if attach:
            attach_file = open(attach, 'rb')
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attach_file.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename= {attach}')
            msg.attach(part)

        server.sendmail(from_email, to, msg.as_string())
        print("Sent")
    except smtplib.SMTPServerDisconnected as e:
        print(f"SMTP Server Disconnected: {e}")


def update_status(worksheet, log_sheet, row, status):
    worksheet.cell(row=row, column=4, value=status)
    worksheet.cell(row=row, column=5, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    log_sheet.append([row, status, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])


def main(email, pwd, spreadsheet):
    server = config_service(email, pwd)

    # Read the spreadsheet
    workbook = openpyxl.load_workbook(spreadsheet)
    worksheet = workbook.active

    # Create or get the log sheet
    if 'Log' not in workbook.sheetnames:
        log_sheet = workbook.create_sheet('Log')
        log_sheet.append(['Row', 'Status', 'Timestamp'])
    else:
        log_sheet = workbook['Log']

    for row in range(2, worksheet.max_row + 1):  # Starting from the second row, assuming the first row is the header
        recipient = worksheet.cell(row=row, column=1).value
        message = worksheet.cell(row=row, column=2).value
        attachment = worksheet.cell(row=row, column=3).value
        status = worksheet.cell(row=row, column=4).value

        if status == 'waiting':
            send_mail(server, email, recipient, "Email Subject", message, attachment)
            update_status(worksheet, log_sheet, row, 'sent')

    workbook.save(spreadsheet)
    server.quit()
    print("Spreadsheet checked and updated. Exiting...")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('email', help='Email login credentials')
    parser.add_argument('pwd', help='Email password')
    parser.add_argument('spreadsheet', help='Path to the spreadsheet to be checked and updated')
    args = parser.parse_args()

    main(args.email, args.pwd, args.spreadsheet)
